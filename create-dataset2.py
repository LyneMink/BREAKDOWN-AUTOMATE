# ============================================================================
# construire_dataset.py
#
# Dataset final équilibré — Option A : 90 exemples par catégorie
# Split : 70 train / 10 test / 10 eval (par catégorie)
# Total : 1470 train / 210 test / 210 eval = 1890 exemples
# ============================================================================

import os
import re
import json
import random
import openpyxl
from collections import defaultdict, Counter

random.seed(42)

# ── Chemins ──────────────────────────────────────────────────────────────────
RAW_DIR    = "/home/minkoh-som/Pictures/Holy-Finetunning/dataset/raw"
OUTPUT_DIR = "/home/minkoh-som/Pictures/Holy-Finetunning/dataset/dataset_final"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Quota fixe par catégorie ──────────────────────────────────────────────────
QUOTA_TOTAL = 90   # exemples réels max par catégorie
Q_TRAIN     = 70   # par catégorie
Q_TEST      = 10   # par catégorie
Q_EVAL      = 10   # par catégorie

# ── Correspondance fichier → catégorie ───────────────────────────────────────
FICHIERS_CATEGORIES = {
    "power_defaut_ge.xlsx"                       : "Defaut GE & Power Cabinet",
    "power_aktivco_ge_fixed (3).xlsx"            : "AKTIVCO Defaut GE & Power Cabinet",
    "power_coupure_eneo_v3.xlsx"                 : "Coupure ENEO & Baisse de tension",
    "power_aktivco_eneo.xlsx"                    : "AKTIVCO Coupure ENEO & Baisse de tension",
    "sharing_dataset.xlsx"                       : "Sharing",
    "sites_strategiques_datacenter.xlsx"         : "Sites strategiques & DataCenter",
    "sites_strategiques_mll_pylone_dataset.xlsx" : "Sites strategiques, MLL, Pylone, etc.",
    "ocm_huawei_v2.xlsx"                         : "Projet OCM (HUAWEI)",
    "ocm_zte_nokia_corrected.xlsx"               : "Projet OCM (ZTE, NOKIA, autres projets)",
    "bss_hardware_issue.xlsx"                    : "BSS Hardware issue",
    "access_issue_dataset.xlsx"                  : "ACCESS-ISSUE",
    "mpr_issue_dataset.xlsx"                     : "MPR issue",
    "odu_hs_dataset_final.xlsx"                  : "ODU HS",
    "ip_vlan_dataset.xlsx"                       : "IP & VLAN",
    "fiber_aof_dataset.xlsx"                     : "fiber AOF",
    "fiber_camtel_dataset.xlsx"                  : "fiber CAMTEL",
    "spare_issue_dataset.xlsx"                   : "SPARE-ISSUE",
    "spare_hs_dataset.xlsx"                      : "SPARE-HS",
    "sat_dataset.xlsx"                           : "SAT",
    "exclu_dataset_final.xlsx"                   : "EXCLU",
    "warehouse_huawei_dataset.xlsx"              : "Warehouse HUAWEI",
}

# ── System Prompt ─────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """Tu es un expert en analyse d'incidents réseau télécom au monde.
Tu reçois un commentaire d'incident et tu dois identifier la cause racine
parmi les 21 catégories exactes ci-dessous.

IMPORTANT — ÉQUIPROBABILITÉ DES CATÉGORIES :
Toutes les 21 catégories sont également probables.
Tu ne dois favoriser aucune catégorie par défaut.
Ta classification repose UNIQUEMENT sur les règles métier ci-dessous
et le contenu exact du commentaire, Owner, Topology et Vendors reçus.
Ne jamais supposer une catégorie sans justification explicite dans les données.

Selon le type d'incident tu peux aussi recevoir :
- Owner    : le gestionnaire du site
- Topology : la configuration électrique du site
- Vendors  : l'équipementier radio (ZTE, HUAWEI, NOKIA)

RÈGLES MÉTIER — À APPLIQUER DANS L'ORDRE DE PRIORITÉ

RÈGLE 1 — OWNER MTN :
Si Owner = MTN ou MTN-Sharing ou MTN Sharing
→ Cause : Sharing
→ La Topology n'est pas nécessaire pour cette règle.

RÈGLE 2 — OWNER OCM :
Si Owner = OCM ou OCM-Strategie ou OCM/strategique ou OCM/datacenter
→ Cause : Sites strategiques & DataCenter
→ La Topology n'est pas nécessaire pour cette règle.

RÈGLE 3 — PROJET OCM RADIO (VENDORS) :
Si le commentaire indique des travaux de projet OCM sur les équipements radio :
  Vendors = HUAWEI          → Projet OCM (HUAWEI)
  Vendors = ZTE ou NOKIA    → Projet OCM (ZTE, NOKIA, autres projets)
  Vendors non précisé       → Projet OCM (ZTE, NOKIA, autres projets)

RÈGLE 4 — TOPOLOGY (source d'énergie) :
Topologies AVEC source alternative (GE, Solar, Hybrid, Lithium...) :
Bad Grid + GE, Bad Grid + Solar + GE, Bad-Grid-GE, DG+Grid, GE, Bad Grid, Gen Only, Gen-, Gen-Hybrid, Gen-Lithium, Gen-Solar AER, GenOnly, Gird-Hybrid-Gen, Good Grid + GE, Good Grid + Solar + GE, Good-Grid + solar + SE, Grid + Gen, Grid + Solar, Grid - Gen, Grid Gen, Grid(fault) Gen, Grid-Gen, Grid-Gen (LL), Grid-Gen-Lithium, Grid-Gen-Solar, Grid-Genlithium, Grid-Hybrid-Gen, Grid-Lithium, Grid-Solar, Grid-Solar-Gen, GRID GEN, HGB, Hibrid solar - s3, Hybrid, Hybrid Gen, Hybrid Solar - S1, Hybrid Solar - S2, Hybrid Solar - S3, Hybrid Solar 3, Hybrid-Gen, Hybrid-Gen-Lithium, Lithium, medium Grid GE, Medium Grid + GE, Medium Grid + Solar + GE, Medium grid solar + GE, meduim grid + GE, Pure Solar, Solar, Solar AER-Gen-Lithium, Solar Gen, Solar Only, Solar+ Gen, Solar-Gen, Solar-Gen-Lithium, Solar-Grid-Gen, Solar-Grid-Gen-Lithium, Solar-Hybrid-Gen, SolarAER-Gen-Lithium

Topologies SANS source alternative (dépend uniquement d'ENEO) :
good grid, good grid 100%, good grid no ge, good grid no ge - 8h, good grid no ge - 12h, good grid no ge 8h, grid only, grid-only, gridonly

RÈGLE 5 — OWNER + TOPOLOGY → CAUSE ÉNERGIE :
Owner = CAMUSAT ou ESCO :
  + Topology AVEC source alternative  → AKTIVCO Defaut GE & Power Cabinet
  + Topology SANS source alternative  → AKTIVCO Coupure ENEO & Baisse de tension
Owner = IHS ou I_HS ou HIS :
  + Topology AVEC source alternative  → Defaut GE & Power Cabinet
  + Topology SANS source alternative  → Coupure ENEO & Baisse de tension

LISTE DES 21 CATÉGORIES EXACTES :
1.  Coupure ENEO & Baisse de tension
2.  AKTIVCO Coupure ENEO & Baisse de tension
3.  Defaut GE & Power Cabinet
4.  AKTIVCO Defaut GE & Power Cabinet
5.  Sharing
6.  Sites strategiques & DataCenter
7.  Sites strategiques, MLL, Pylone, etc.
8.  Projet OCM (HUAWEI)
9.  Projet OCM (ZTE, NOKIA, autres projets)
10. BSS Hardware issue
11. ACCESS-ISSUE
12. MPR issue
13. ODU HS
14. IP & VLAN
15. fiber AOF
16. fiber CAMTEL
17. SPARE-ISSUE
18. SPARE-HS
19. SAT
20. EXCLU
21. Warehouse HUAWEI

FORMAT DE RÉPONSE — TOUJOURS CE FORMAT EXACT :
Analyse : [2-3 lignes sur la situation, incident clôturé ou en cours]
Justification : [règle appliquée + pourquoi pas les autres causes]
Cause : [libellé exact parmi les 21 catégories ci-dessus]"""

# ── Topologies ────────────────────────────────────────────────────────────────
TOPOLOGIES_AVEC_SOURCE = [
    "bad grid + ge", "bad grid + solar + ge", "bad-grid-ge", "dg+grid",
    "ge", "bad grid", "gen only", "gen-", "gen-hybrid", "gen-lithium",
    "gen-solar aer", "genonly", "gird-hybrid-gen", "good grid + ge",
    "good grid + solar + ge", "good-grid + solar + se", "grid + gen",
    "grid + solar", "grid - gen", "grid gen", "grid(fault) gen", "grid-gen",
    "grid-gen (ll)", "grid-gen-lithium", "grid-gen-solar", "grid-genlithium",
    "grid-hybrid-gen", "grid-lithium", "grid-solar", "grid-solar-gen",
    "grid gen", "hgb", "hibrid solar - s3", "hybrid", "hybrid gen",
    "hybrid solar - s1", "hybrid solar - s2", "hybrid solar - s3",
    "hybrid solar 3", "hybrid-gen", "hybrid-gen-lithium", "lithium",
    "medium grid ge", "medium grid + ge", "medium grid + solar + ge",
    "medium grid solar + ge", "meduim grid + ge", "pure solar", "solar",
    "solar aer-gen-lithium", "solar gen", "solar only", "solar+ gen",
    "solar-gen", "solar-gen-lithium", "solar-grid-gen",
    "solar-grid-gen-lithium", "solar-hybrid-gen", "solaraer-gen-lithium"
]

TOPOLOGIES_SANS_SOURCE = [
    "good grid", "good grid 100%", "good grid no ge",
    "good grid no ge - 8h", "good grid no ge - 12h",
    "good grid no ge 8h", "grid only", "grid-only", "gridonly"
]

def a_source_alternative(topology):
    t = topology.lower().strip()
    for s in TOPOLOGIES_SANS_SOURCE:
        if s == t:
            return False
    for s in TOPOLOGIES_AVEC_SOURCE:
        if s in t:
            return True
    if any(k in t for k in ["ge", "solar", "hybrid", "gen", "lithium"]):
        return True
    return None

def detecter_statut(commentaire):
    com_lower = commentaire.lower()
    mots_cloture = [
        "site up", "restored", "clôturé", "cloture", "resolved",
        "normalized", "normalisé", "power restored", "site stable",
        "closed", "intervention done", "done", "completed", "finalisé"
    ]
    if any(m in com_lower for m in mots_cloture):
        return "clôturé"
    return "en cours"

def extraire_champs_user(user_content):
    commentaire = ""
    owner       = ""
    topology    = ""
    vendors     = ""
    lignes = user_content.split("\n")
    for ligne in lignes:
        l = ligne.strip()
        if l.lower().startswith("owner :"):
            owner = l.split(":", 1)[-1].strip()
        elif l.lower().startswith("topology :") or l.lower().startswith("topolopgy :"):
            topology = l.split(":", 1)[-1].strip()
        elif l.lower().startswith("vendors :"):
            vendors = l.split(":", 1)[-1].strip()
        elif l.lower().startswith("commentaire :"):
            commentaire = l.split(":", 1)[-1].strip().strip('"')
    if not commentaire:
        commentaire = user_content.strip()
    return commentaire, owner, topology, vendors

def generer_analyse(categorie, commentaire, owner, topology, vendors):
    statut = detecter_statut(commentaire)
    templates = {
        "Defaut GE & Power Cabinet": (
            f"Le site présente une défaillance sur le générateur ou le power cabinet "
            f"impactant la stabilité énergétique. Les équipes FME et power poursuivent "
            f"les investigations et actions correctives sur les équipements power. "
            f"L'incident est {statut}."
        ),
        "AKTIVCO Defaut GE & Power Cabinet": (
            f"Le commentaire montre une anomalie énergétique liée au générateur ou au "
            f"power cabinet affectant la stabilité du site. Les équipes techniques "
            f"poursuivent les investigations et les actions correctives sur les "
            f"équipements power. L'incident est {statut}."
        ),
        "Coupure ENEO & Baisse de tension": (
            f"Le site subit une panne ou instabilité du réseau commercial ENEO. "
            f"Le site dépend uniquement du réseau électrique public et les équipes "
            f"terrain attendent la normalisation ENEO. L'incident est {statut}."
        ),
        "AKTIVCO Coupure ENEO & Baisse de tension": (
            f"Le site subit une instabilité du réseau commercial ENEO avec plusieurs "
            f"alarmes power actives. Les équipes terrain poursuivent le suivi "
            f"opérationnel et les investigations power. L'incident est {statut}."
        ),
        "Sharing": (
            f"Le site présente des alarmes power sur une infrastructure MTN sharing. "
            f"Les équipes MTN sont activées pour investigation et actions correctives "
            f"sur l'alimentation électrique. L'incident est {statut}."
        ),
        "Sites strategiques & DataCenter": (
            f"Le commentaire révèle une anomalie critique sur une infrastructure "
            f"stratégique ou DataCenter. Les équipes OCM/Stratégique poursuivent "
            f"les actions de stabilisation et sécurisation. L'incident est {statut}."
        ),
        "Sites strategiques, MLL, Pylone, etc.": (
            f"Le commentaire révèle une anomalie environnementale ou structurelle "
            f"impactant le site. Les interventions sont limitées par des conditions "
            f"climatiques, physiques ou environnementales. L'incident est {statut}."
        ),
        "Projet OCM (HUAWEI)": (
            f"Le commentaire décrit des travaux de migration ou commissioning sur "
            f"les équipements radio HUAWEI. Le site est temporairement hors service "
            f"pendant l'activité projet OCM. L'incident est {statut}."
        ),
        "Projet OCM (ZTE, NOKIA, autres projets)": (
            f"Le commentaire décrit des travaux de migration ou commissioning sur "
            f"les équipements radio ZTE ou NOKIA. Le site est temporairement hors "
            f"service pendant l'activité projet OCM. L'incident est {statut}."
        ),
        "BSS Hardware issue": (
            f"Le commentaire révèle une défaillance matérielle radio impactant les "
            f"équipements du site. Les équipes terrain poursuivent les investigations "
            f"et opérations de remplacement des modules radio. L'incident est {statut}."
        ),
        "ACCESS-ISSUE": (
            f"Le commentaire révèle un problème d'accès au site empêchant les équipes "
            f"techniques d'effectuer les investigations ou actions correctives. "
            f"Les interventions sont bloquées en attente d'autorisation ou de "
            f"sécurisation de l'accès. L'incident est {statut}."
        ),
        "MPR issue": (
            f"Le commentaire présente une panne de transmission MPR impactant "
            f"directement les services et la supervision du site. Les investigations "
            f"terrain et opérations de remplacement des équipements MPR sont en cours. "
            f"L'incident est {statut}."
        ),
        "ODU HS": (
            f"Le lien de transmission est perturbé à cause d'une anomalie matérielle "
            f"ou logicielle sur l'ODU. Les équipes terrain effectuent des vérifications "
            f"locales et des opérations de remplacement sur la chaîne antennaire. "
            f"L'incident est {statut}."
        ),
        "IP & VLAN": (
            f"Le commentaire montre une dégradation des services liée au routage IP "
            f"ou aux VLANs. Les investigations se concentrent sur la connectivité "
            f"IPRAN, les interfaces S1 et les configurations réseau. "
            f"L'incident est {statut}."
        ),
        "fiber AOF": (
            f"Le commentaire confirme une anomalie sur la liaison fibre optique "
            f"BBU/RRU affectant la transmission du site. Les équipes terrain "
            f"poursuivent les investigations sur la fibre, les SFP et la connectique. "
            f"L'incident est {statut}."
        ),
        "fiber CAMTEL": (
            f"Le commentaire révèle une panne backbone CAMTEL affectant la "
            f"connectivité des sites. Les équipes CAMTEL poursuivent les "
            f"investigations et actions de restauration du backbone. "
            f"L'incident est {statut}."
        ),
        "SPARE-ISSUE": (
            f"Le commentaire présente un problème de disponibilité de spare "
            f"empêchant la restauration du site. Les équipes poursuivent les "
            f"relances et procédures de commande pour permettre l'intervention. "
            f"L'incident est {statut}."
        ),
        "SPARE-HS": (
            f"Le commentaire révèle un spare reçu mais défectueux empêchant "
            f"la restauration du site. Les équipes poursuivent les vérifications "
            f"et nouvelles demandes de remplacement. L'incident est {statut}."
        ),
        "SAT": (
            f"Le commentaire révèle une instabilité de transmission satellite VSAT "
            f"avec plusieurs alarmes et perturbations. Les investigations se "
            f"concentrent sur les modems, BUC/LNB et qualité du signal satellite. "
            f"L'incident est {statut}."
        ),
        "EXCLU": (
            f"Le commentaire présente un contexte d'insécurité ou de vandalisme "
            f"bloquant toute opération sur le site. Les actions sont suspendues "
            f"en attente de sécurisation du site. L'incident est {statut}."
        ),
        "Warehouse HUAWEI": (
            f"Le commentaire présente un workflow logistique Huawei avec transport "
            f"ou expédition de spare vers le site. Les équipes Warehouse et FME "
            f"poursuivent les activités logistiques et la réception des équipements. "
            f"L'incident est {statut}."
        ),
    }
    return templates.get(categorie, f"Incident en cours sur le site. L'analyse est {statut}.")

def generer_justification(categorie, commentaire, owner, topology, vendors):
    if categorie == "Sharing":
        return (
            f"Owner = {owner}. La Règle 1 s'applique directement : "
            f"Owner MTN/MTN-Sharing → Cause Sharing, la Topology n'est pas "
            f"nécessaire pour cette règle. Ce n'est pas Coupure ENEO ni Defaut GE "
            f"car la gestion énergétique est assurée par MTN sharing et non par "
            f"les équipes IHS/CAMUSAT/ESCO."
        )
    if categorie == "Sites strategiques & DataCenter":
        return (
            f"Owner = {owner}. La Règle 2 s'applique directement : "
            f"Owner OCM/OCM-Strategie/OCM-Datacenter → Cause Sites strategiques "
            f"& DataCenter, la Topology n'est pas nécessaire pour cette règle. "
            f"Ce n'est pas une catégorie énergie classique car ce site est géré "
            f"par l'équipe OCM/Stratégique avec des contraintes spécifiques."
        )
    if categorie == "Projet OCM (HUAWEI)":
        return (
            f"Vendors = {vendors} et le commentaire décrit explicitement des "
            f"opérations OCM radio (swap, migration, commissioning, upgrade, "
            f"intégration réseau). La Règle 3 s'applique : Vendors HUAWEI → "
            f"Projet OCM (HUAWEI). Ce n'est pas Projet OCM ZTE/NOKIA car Vendors "
            f"est HUAWEI. Ce n'est pas BSS Hardware issue car il s'agit d'une "
            f"activité projet planifiée et non d'une panne matérielle."
        )
    if categorie == "Projet OCM (ZTE, NOKIA, autres projets)":
        vendor_str = vendors if vendors else "ZTE/NOKIA"
        return (
            f"Vendors = {vendor_str} et le commentaire décrit explicitement des "
            f"opérations OCM radio (swap, migration, commissioning, upgrade). "
            f"La Règle 3 s'applique : Vendors ZTE/NOKIA → Projet OCM "
            f"(ZTE, NOKIA, autres projets). Ce n'est pas Projet OCM HUAWEI car "
            f"Vendors est {vendor_str}. Ce n'est pas BSS Hardware issue car "
            f"il s'agit d'une activité projet planifiée et non d'une panne."
        )
    cats_energie = [
        "Defaut GE & Power Cabinet",
        "AKTIVCO Defaut GE & Power Cabinet",
        "Coupure ENEO & Baisse de tension",
        "AKTIVCO Coupure ENEO & Baisse de tension",
    ]
    if categorie in cats_energie and owner and topology:
        source = a_source_alternative(topology)
        source_str = (
            "AVEC source alternative" if source
            else "SANS source alternative" if source is False
            else "dont la source est à analyser"
        )
        source_expl = (
            f"le site dispose d'un générateur ou d'une source hybride capable "
            f"de produire de l'électricité indépendamment du réseau ENEO. "
            f"Le problème concerne donc le générateur ou le power cabinet."
            if source else
            f"le site dépend uniquement du réseau ENEO, il n'y a pas de générateur "
            f"ni de source alternative. Toute panne est donc une coupure ou baisse "
            f"de tension ENEO."
        )
        if categorie == "AKTIVCO Defaut GE & Power Cabinet":
            return (
                f"Owner = {owner} et Topology = {topology}. {topology} appartient "
                f"à la liste des topologies {source_str} : {source_expl} "
                f"Ce n'est pas Defaut GE & Power Cabinet car Owner est {owner} "
                f"(CAMUSAT/ESCO) et non IHS/HIS. Ce n'est pas AKTIVCO Coupure ENEO "
                f"car la topology possède une source alternative."
            )
        elif categorie == "AKTIVCO Coupure ENEO & Baisse de tension":
            return (
                f"Owner = {owner} et Topology = {topology}. {topology} appartient "
                f"à la liste des topologies {source_str} : {source_expl} "
                f"Ce n'est pas Coupure ENEO & Baisse de tension car Owner est "
                f"{owner} (CAMUSAT/ESCO) et non IHS/HIS. Ce n'est pas AKTIVCO "
                f"Defaut GE car il n'y a pas de source alternative sur ce site."
            )
        elif categorie == "Defaut GE & Power Cabinet":
            return (
                f"Owner = {owner} et Topology = {topology}. {topology} appartient "
                f"à la liste des topologies {source_str} : {source_expl} "
                f"Ce n'est pas AKTIVCO Defaut GE & Power Cabinet car Owner est "
                f"{owner} (IHS/HIS) et non CAMUSAT/ESCO. Ce n'est pas Coupure ENEO "
                f"car la topology possède une source alternative."
            )
        elif categorie == "Coupure ENEO & Baisse de tension":
            return (
                f"Owner = {owner} et Topology = {topology}. {topology} appartient "
                f"à la liste des topologies {source_str} : {source_expl} "
                f"Ce n'est pas AKTIVCO Coupure ENEO & Baisse de tension car Owner "
                f"est {owner} (IHS/HIS) et non CAMUSAT/ESCO. Ce n'est pas Defaut GE "
                f"car il n'y a pas de source alternative sur ce site."
            )
    justifs_autres = {
        "BSS Hardware issue": (
            "Les éléments du commentaire concernent explicitement des modules "
            "radio/BSS (FRGU, FXDB, FRMF, ARGA, ABIA, RRU) avec des actions "
            "typiques telles que remplacement, swap, reset ou check interconnexion. "
            "Ce n'est pas MPR issue car les équipements concernés sont des modules "
            "radio et non des équipements transmission. Ce n'est pas ODU HS car "
            "il n'y a pas de problème sur la chaîne antennaire ODU."
        ),
        "ACCESS-ISSUE": (
            "Les éléments du commentaire concernent explicitement des problèmes "
            "d'accès terrain (zone rouge, ghost town, accès refusé, négociations, "
            "sécurité). Même si un problème technique existe, la cause principale "
            "est l'impossibilité d'accéder au site. Ce n'est pas EXCLU car il "
            "n'y a pas de vandalisme ou d'insécurité totale."
        ),
        "MPR issue": (
            "Les éléments du commentaire concernent explicitement des équipements "
            "transmission MPR (chassis MSS8, carte EAC, Core EVO, MPT, P8ETH) "
            "avec des actions typiques telles que remplacement ou check MPR status. "
            "Ce n'est pas BSS Hardware issue car les équipements sont des "
            "équipements transmission et non des modules radio."
        ),
        "ODU HS": (
            "Les éléments du commentaire concernent explicitement l'ODU, la "
            "chaîne antennaire ou des problèmes de transmission microwave. "
            "Ce n'est pas MPR issue car le problème est sur l'ODU et non sur "
            "les équipements MPR internes."
        ),
        "IP & VLAN": (
            "Les éléments du commentaire concernent principalement des problèmes "
            "IP, VLAN, routage, interfaces réseau ou instabilités IPRAN. "
            "Ce n'est pas fiber AOF car il ne s'agit pas d'un problème physique "
            "de fibre optique."
        ),
        "fiber AOF": (
            "Les éléments du commentaire concernent principalement des problèmes "
            "de fibre optique locale, SFP, liaison BBU-RRU ou connectique FO. "
            "Ce n'est pas fiber CAMTEL car il s'agit d'une fibre locale "
            "et non du backbone CAMTEL."
        ),
        "fiber CAMTEL": (
            "Les éléments du commentaire concernent principalement des incidents "
            "liés au backbone CAMTEL avec escalades et tickets CAMTEL. "
            "Ce n'est pas fiber AOF car il ne s'agit pas d'une fibre locale "
            "BBU-RRU mais du backbone opérateur CAMTEL."
        ),
        "SPARE-ISSUE": (
            "Les éléments du commentaire concernent principalement des problèmes "
            "de disponibilité de spare, commande HW SPMS ou attente "
            "d'approvisionnement. Ce n'est pas SPARE-HS car le spare n'a pas "
            "encore été reçu et testé défectueux."
        ),
        "SPARE-HS": (
            "Les éléments du commentaire concernent explicitement des spares "
            "déjà reçus mais déclarés faulty ou défectueux après vérification. "
            "Ce n'est pas SPARE-ISSUE car le spare a été reçu mais il est "
            "défectueux."
        ),
        "SAT": (
            "Les éléments du commentaire concernent principalement des équipements "
            "VSAT, problèmes de signal satellite ou défauts modem/BUC/LNB. "
            "Ce n'est pas fiber AOF car il s'agit d'une transmission satellite "
            "et non d'une fibre optique."
        ),
        "EXCLU": (
            "Les éléments du commentaire concernent explicitement un environnement "
            "non sécurisé (vandalisme, zone rouge, ghost town, équipements pillés). "
            "Ce n'est pas ACCESS-ISSUE car il ne s'agit pas d'un simple problème "
            "d'accès temporaire mais d'une exclusion totale pour raisons sécuritaires."
        ),
        "Warehouse HUAWEI": (
            "Les éléments du commentaire concernent explicitement des workflows "
            "de transport ou livraison de matériels Huawei (ETD/ETA, WH, HW SPMS). "
            "Ce n'est pas SPARE-ISSUE car le spare est en transit ou en cours "
            "de livraison."
        ),
        "Sites strategiques, MLL, Pylone, etc.": (
            "Les éléments du commentaire concernent explicitement des facteurs "
            "environnementaux (pluie, foudre, infiltrations, pylône, MLL) ou "
            "des contraintes physiques structurelles. Ce n'est pas ACCESS-ISSUE "
            "car le problème est environnemental et non un problème d'accès."
        ),
    }
    return justifs_autres.get(
        categorie,
        f"Les éléments du commentaire correspondent à la catégorie {categorie}."
    )

def construire_exemple(user_content, categorie):
    commentaire, owner, topology, vendors = extraire_champs_user(user_content)
    analyse       = generer_analyse(categorie, commentaire, owner, topology, vendors)
    justification = generer_justification(categorie, commentaire, owner, topology, vendors)
    assistant_content = (
        f"Analyse : {analyse}\n\n"
        f"Justification : {justification}\n\n"
        f"Cause : {categorie}"
    )
    return {
        "messages": [
            {"role": "system",    "content": SYSTEM_PROMPT},
            {"role": "user",      "content": user_content},
            {"role": "assistant", "content": assistant_content},
        ]
    }

def lire_fichier_excel(chemin):
    wb = openpyxl.load_workbook(chemin)
    ws = wb.active
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).lower().strip()] = cell.column - 1
    col_user = None
    for nom in ["user", "d", "commentaire"]:
        if nom in headers:
            col_user = headers[nom]
            break
    if col_user is None:
        col_user = 3
    user_contents = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_user < len(row) and row[col_user]:
            val = str(row[col_user]).strip()
            if val and val.lower() not in ["nan", "none", ""]:
                user_contents.append(val)
    return user_contents

# ── Programme principal ───────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("CONSTRUCTION DU DATASET — Option A (90 par catégorie)")
    print(f"Split : {Q_TRAIN} train / {Q_TEST} test / {Q_EVAL} eval")
    print("=" * 60)

    train_all = []
    test_all  = []
    eval_all  = []

    for fichier, categorie in FICHIERS_CATEGORIES.items():
        chemin = os.path.join(RAW_DIR, fichier)
        if not os.path.exists(chemin):
            print(f"  ❌ MANQUANT : {fichier}")
            continue

        user_contents = lire_fichier_excel(chemin)
        dispo = len(user_contents)

        if dispo < QUOTA_TOTAL:
            print(f"  ⚠️  {categorie:<45} : {dispo} dispo < {QUOTA_TOTAL} requis — IGNORÉ")
            continue

        # Mélanger et prendre exactement 90
        random.shuffle(user_contents)
        selectionnes = user_contents[:QUOTA_TOTAL]

        # Construire les exemples
        exemples = [construire_exemple(uc, categorie) for uc in selectionnes]

        # Découper
        train_all.extend(exemples[:Q_TRAIN])
        test_all.extend(exemples[Q_TRAIN:Q_TRAIN + Q_TEST])
        eval_all.extend(exemples[Q_TRAIN + Q_TEST:Q_TRAIN + Q_TEST + Q_EVAL])

        print(f"  ✅ {categorie:<45} : {dispo} dispo → {Q_TRAIN} train / {Q_TEST} test / {Q_EVAL} eval")

    # Mélanger les fichiers finaux
    random.shuffle(train_all)
    random.shuffle(test_all)
    random.shuffle(eval_all)

    # Sauvegarder
    def sauvegarder(exemples, nom):
        chemin = os.path.join(OUTPUT_DIR, nom)
        with open(chemin, "w", encoding="utf-8") as f:
            for ex in exemples:
                f.write(json.dumps(ex, ensure_ascii=False) + "\n")
        print(f"  💾 {nom} → {len(exemples)} exemples")

    print(f"\n{'='*60}")
    print("SAUVEGARDE")
    print(f"{'='*60}")
    sauvegarder(train_all, "train.jsonl")
    sauvegarder(test_all,  "test.jsonl")
    sauvegarder(eval_all,  "eval.jsonl")

    # Vérification distribution train
    print(f"\n{'='*60}")
    print("VÉRIFICATION DISTRIBUTION TRAIN")
    print(f"{'='*60}")
    dist = Counter()
    for ex in train_all:
        for msg in ex["messages"]:
            if msg["role"] == "assistant" and "Cause :" in msg["content"]:
                cause = msg["content"].split("Cause :")[-1].strip().split("\n")[0].strip()
                dist[cause] += 1

    for cat, n in sorted(dist.items(), key=lambda x: -x[1]):
        statut = "✅" if n == Q_TRAIN else "⚠️"
        print(f"    {statut} [{n:3d}/{Q_TRAIN}] {cat}")

    print(f"\n  Train : {len(train_all)} exemples")
    print(f"  Test  : {len(test_all)} exemples")
    print(f"  Eval  : {len(eval_all)} exemples")
    print(f"  Total : {len(train_all)+len(test_all)+len(eval_all)} exemples")
    print(f"\n✅ Dataset généré dans : {OUTPUT_DIR}")

if __name__ == "__main__":
    main()